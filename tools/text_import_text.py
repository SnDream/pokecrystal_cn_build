#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl import load_workbook
import subprocess

wb = load_workbook('./text.xlsx', data_only=True)

class textblock:
    def __init__(self):
        self.dmap = ''
        self.omap = ''
        self.dlabel = ''
        self.olabel = ''
        self.ofname = ''
        self.oeomjp = ''
        self.oeomen = ''
        self.next_CR = False
        self.enlist = []
        self.jplist = []
        self.cnlist = []
        self.hint = []
        self.comment = []
        self.ctrl = []
        self.asm = ''


def get_textdata():
    ws = wb['标']
    tb_dict = dict()
    for wbi in range(2, ws.max_row):
        tb = textblock()
        if ws.cell(row = wbi, column = 1).value is not None:
            tb.dmap = ws.cell(row = wbi, column = 1).value
        if ws.cell(row = wbi, column = 2).value is not None:
            tb.omap = ws.cell(row = wbi, column = 2).value
        if ws.cell(row = wbi, column = 3).value is not None:
            tb.dlabel = ws.cell(row = wbi, column = 3).value
        tb.olabel = ws.cell(row = wbi, column = 4).value
        tb.ofname = ws.cell(row = wbi, column = 5).value
        tb.oeomjp = ws.cell(row = wbi, column = 6).value
        tb.oeomen = ws.cell(row = wbi, column = 7).value
        tb_dict[tb.olabel] = tb
    for i in range(9):
        ws = wb['文' + str(i + 1)]
        enlist = []
        jplist = []
        cnlist = []
        hint = []
        comment = []
        ctrl = []
        olabel = ''
        for wbi in range(1, ws.max_row + 1):
            cen = ws.cell(row = wbi, column = 1).value
            if cen is None: cen = ''
            cjp = ws.cell(row = wbi, column = 3).value
            if cjp is None: cjp = ''
            ccn = ws.cell(row = wbi, column = 5).value
            if ccn is None: ccn = ''
            cht = ws.cell(row = wbi, column = 8).value
            if cht is None: cht = ''
            ccm = ws.cell(row = wbi, column = 9).value
            if ccm is None: ccm = ''
            ctr = ws.cell(row = wbi, column = 10).value
            if ctr is None: ctr = ''
            if '英文' in cen or '结束' in cen:
                if olabel != '':
                    enlist_end_cnt = 0
                    jplist_end_cnt = 0
                    while len(enlist) > 0 and enlist[-1] == '': 
                        enlist.pop()
                        enlist_end_cnt += 1
                    while len(jplist) > 0 and jplist[-1] == '': 
                        jplist.pop()
                        jplist_end_cnt += 1
                    while len(cnlist) > 0 and cnlist[-1] == '': cnlist.pop()
                    if enlist_end_cnt > 2 and jplist_end_cnt > 2:
                        print('END WITH EMPTY PARA FOUND! ', olabel)
                    if len(ctrl) > 0 and ctrl[0] == 'LINE_CR':
                        tb_dict[olabel].next_CR = True
                        ctrl.pop(0)
                    # if olabel == 'anata_msg_000_Kojindat':
                    #     ret1,codegit = subprocess.getstatusoutput('git -C ./pokecrystal_cn/ rev-parse --short HEAD')
                    #     ret2,textgit = subprocess.getstatusoutput('git -C . rev-parse --short HEAD')
                    #     print('INSERT GIT INFORMATION', ret1, codegit, ret2, textgit)
                    #     cnlist.insert(0, '感谢您参与《精灵宝可梦')
                    #     cnlist.insert(1, '水晶版》汉化版的测试！')
                    #     cnlist.insert(2, '')
                    #     cnlist.insert(3, '请在报告问题时')
                    #     cnlist.insert(4, '提供下一页的信息：')
                    #     cnlist.insert(5, '')
                    #     cnlist.insert(6, '代码：'+codegit)
                    #     cnlist.insert(7, '文本：'+textgit)
                    #     cnlist.insert(8, '')
                    #     cnlist.insert(9, '最后，请勿外传此测试')
                    #     cnlist.insert(10, 'ROM！下面开始游戏……')
                    #     cnlist.insert(11, '')
                    tb_dict[olabel].enlist = enlist
                    tb_dict[olabel].jplist = jplist
                    tb_dict[olabel].cnlist = cnlist
                    tb_dict[olabel].hint = hint
                    tb_dict[olabel].comment = comment
                    tb_dict[olabel].ctrl = ctrl
                    for hinttoken in hint:
                        if hinttoken == '仁': continue
                        if hinttoken == '真': continue
                        hintpass = False
                        for cnline in cnlist:
                            if hinttoken in cnline:
                                hintpass = True
                                break
                        if hintpass : continue
                        print(olabel + ' HINT LOSS : ' + hinttoken)
                        print(''.join(cnlist))
                        
                    for jpline in jplist:
                        if '<USER>' in jpline:
                            hintpass = False
                            for cnline in cnlist:
                                if '<USER>' in cnline:
                                    hintpass = True
                                    break
                            if not hintpass :
                                print(olabel + ' HINT LOSS : ' + '<USER>')
                        if '<TARGET>' in jpline:
                            hintpass = False
                            for cnline in cnlist:
                                if '<TARGET>' in cnline:
                                    hintpass = True
                                    break
                            if not hintpass :
                                print(olabel + ' HINT LOSS : ' + '<TARGET>')

                if '英文' in cen:
                    olabel = ctr
                    enlist = []
                    jplist = []
                    cnlist = []
                    hint = []
                    comment = []
                    ctrl = []
            else:
                enlist.append(cen)
                jplist.append(cjp)
                cnlist.append(ccn)
                if cjp != '' and cjp == ccn:
                    print('MAYBE UNTRANSLATE', cjp, olabel)
                if cht != '' :
                    for chttoken in cht.split(' '):
                        hint.append(chttoken.split(':')[1])
                if ccm != '' : comment.append(ccm)
                if ctr != '' : ctrl.append(ctr)
                if ctr == '…' :
                    print('WARN', olabel)
    return tb_dict
        


def get_asmfile_set():
    ws = wb['标']
    asmfile_set = set()
    for wbi in range(2, ws.max_row):
        if ws.cell(row = wbi, column = 1).value is not None:
            asmfile_set.add(ws.cell(row = wbi, column = 1).value)
    return asmfile_set

def get_asmfile_data():
    asmfile_data = dict()
    for asmfile_name in get_asmfile_set():
        with open('./pokecrystal_cn/' + asmfile_name) as f:
            asmfile_data[asmfile_name] = f.readlines()
    return asmfile_data

oddict = {}
spoddict = {}
with open('./tools/text_import_text_odctrl.txt') as f:
    for line in f:
        od, fkname = line.strip('\n').split('\t')
        oddict[od] = fkname

spoddict['monster_m03']    = {'text_ram wEnemyMonNick':  '赫拉克罗斯'}
spoddict['nigeta_m01']     = {'text_ram wEnemyMonNick':  '代欧奇希斯'}
spoddict['torikaeru_m01']  = {'text_ram wEnemyMonNick':  '精灵宝可梦'}
spoddict['nige_ok_m02']    = {'text_ram wStringBuffer1': '烟雾球'}
spoddict['koreijou_m01']    = {'text_ram wStringBuffer2': '命中率'}
spoddict['koreijou_m02']    = {'text_ram wStringBuffer2': '闪避率'}
spoddict['msg_giveitem_01_common'] = {'text_ram wStringBuffer3': '10'}
spoddict['TradeEndMSG_001_Eventsub'] = {'text_ram wMonOrItemNameBuffer': '美佐子', 'text_ram wStringBuffer2': '三合一磁怪'}
spoddict['msg2_4_012_Eventsub'] = {'text_ram wMonOrItemNameBuffer': '三合一磁怪', 'text_ram wStringBuffer2': '三合一磁怪'}
spoddict['msg3_0_013_Eventsub'] = {'text_ram wMonOrItemNameBuffer': '三合一磁怪', 'text_ram wStringBuffer2': '三合一磁怪'}
spoddict['msg3_3_016_Eventsub'] = {'text_ram wMonOrItemNameBuffer': '三合一磁怪', 'text_ram wStringBuffer2': '三合一磁怪'}
spoddict['msg3_4_017_Eventsub'] = {'text_ram wMonOrItemNameBuffer': '三合一磁怪', 'text_ram wStringBuffer2': '三合一磁怪'}
spoddict['BreederExpMSG_049_Eventsub'] = {'text_decimal hMoneyTemp, 3, 6': '9900'}
spoddict['msg_oldman_05_r25r0101'] = {'text_ram wStringBuffer3': '走路草'}
spoddict['msg1_2_070_Eventsub'] = {'text_ram wStringBuffer1': 'NEWPKMNAME'}
spoddict['msg1_9_077_Eventsub'] = {'text_ram wStringBuffer1': 'NEWPKMNAME'}
spoddict['msg_017_Monstool'] = {'text_ram wStringBuffer1': '999.9'}
spoddict['msg_zukan_10_1_018_Net_pc'] = {'text_ram wStringBuffer3': '251', 'text_ram wStringBuffer4': '251'}
spoddict['msg_telgirl01_type02_01_d22r0101'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_02_d22r0101'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_03_d22r0101'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_01_r27'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_02_r27'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_03_r27'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_d22r0101'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_d22r0101'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_d22r0101'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_mobile_battleroom_next'] = {'text_ram wStringBuffer3': '7'}
spoddict['msg_telgirl01_type02_01_r32'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_02_r32'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_03_r32'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_01_r26'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_02_r26'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_03_r26'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_r32'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r32'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r32'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_r36'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r36'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r36'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_vending_02_t07r0106'] = {'text_ram wStringBuffer3': '才四个字'}
spoddict['msg_telboy01_type02_01_r44'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r44'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r44'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_r38'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r38'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r38'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_vending_02_t25r1006'] = {'text_ram wStringBuffer3': '才四个字'}
spoddict['msg_telgirl01_type02_02_r34'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type04_01_r46'] = {'text_ram wStringBuffer3': '三个字'}
# spoddict['BattleText_EnemyFled'] = {'text_ram wStringBuffer3': 'PKMONSE'}
spoddict['msg_telboy01_type02_01_r39'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r39'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r39'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_r35'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r35'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_r26'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r26'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r26'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type20_01_r30'] = {'text_ram wStringBuffer4': '小拉达'} # 它最长的PM名字，另一个是拉达
spoddict['msg_telboy01_type01_01_r33'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type01_02_r33'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type01_03_r33'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telgirl01_type02_02_r46'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_r31'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r31'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r31'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_d27r0102'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_d27r0102'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_r43'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r43'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r43'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_01_r30'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_02_r30'] = {'text_ram wStringBuffer3': '三个字'}
spoddict['msg_telboy01_type02_03_r30'] = {'text_ram wStringBuffer3': '三个字'}

def length_check(tb):
    for line in tb.cnlist:
        # replace all name
        length_more = 0
        line = line.replace('—','―')
        line = line.replace('·', '・')
        line = line.replace('<……>', '..')
        line = line.replace('…', '.')
        line = line.replace('¥', 'Y')
        line = line.replace('<PLAYER>', 'PLAYERN')
        line = line.replace('<PLAY_G>', 'PLAYERN')
        line = line.replace('<RIVAL>',  'RIVALNM')
        line = line.replace('<USER>', '敌人的ABCDEFG')
        line = line.replace('<TARGET>', '敌人的ABCDEFG')
        line = line.replace('<ENEMY>', '宝可梦训练家 CARINEY')
        line = line.replace('<SCROLL>', '')
        line = line.replace('\'s', 'S')
        for i, od in enumerate(tb.ctrl):
            if od == 'LINE_CR': 
                continue
            if od == 'text_low':
                length_more += 18
            if tb.olabel in spoddict and od in spoddict[tb.olabel]:
                line = line.replace('【' + str(i) + '】', spoddict[tb.olabel][od])
            else:
                line = line.replace('【' + str(i) + '】', oddict[od])
        if '<' in line: length = 180
        else: length = 0
        cnr = False
        for char in line:
            try:
                clen = len(char.encode(encoding='GB2312'))
            except:
                print(char, 'is not suppport', line, tb.dlabel)
                clen = 1
            if clen == 2:
                # chinese
                if not cnr :
                    length += 2
                    cnr = True
                else:
                    length += 1
                    cnr = False
            else:
                length += 1
                cnr = False
        if length > 18 + length_more:
            print(length, tb.olabel, tb.dlabel, line)

def make_asm(tb):
    # asm_mk = tb.dlabel + ':\n'
    asm_mk = ''
    linec = 0
    parac = 0
    for line in tb.cnlist:
        if line == '':
            if linec == 0 and parac == 0:
                pass
            else:
                linec = 0
                parac += 1
                continue
        if linec == 0 and parac == 0:
            asm_mk += '\ttext "' + line + '"\n'
        elif linec == 0 and parac >= 1:
            asm_mk += '\n\tpara "' + line + '"\n'
        elif linec == 1:
            if tb.next_CR:
                asm_mk += '\tnext "' + line + '"\n'
            else:
                asm_mk += '\tline "' + line + '"\n'
        elif linec >= 2:
            if tb.next_CR:
                asm_mk += '\tnext "' + line + '"\n'
            else:
                asm_mk += '\tcont "' + line + '"\n'
        linec += 1
    for ctrli in range(len(tb.ctrl)):
        asm_mk = asm_mk.replace('【'+str(ctrli)+'】', '@"\n\t' + tb.ctrl[ctrli] + '\n\ttext "')
    asm_mk = asm_mk.replace('|', '')
    asm_mk = asm_mk.replace('text ""', "text_start")
    asm_mk = asm_mk.replace('\ttext "@"\n', '')
    if tb.oeomjp == 'EOMeom':
        asm_mk += '\tdone\n\n'
    elif tb.oeomjp == 'EOMwaiteom':
        asm_mk += '\tprompt\n\n'
    elif tb.oeomjp == 'EOM':
        asm_mk += '\ttext_end\n\n'
    elif tb.oeomjp == 'EOM^2':
        # asm_mk += '\ttext_end\n\n'
        asm_mk += '\ttext_end\n\n\ttext_end ; unused\n\n'
    else:
        raise(Exception(tb.oeomjp + tb.olabel))
    # asm_mk = asm_mk.replace('\n\ttext_start\n\tdone\n', '\n\tdone\n')
    # asm_mk = asm_mk.replace('\n\ttext_start\n\tprompt\n', '\n\tprompt\n')
    asm_mk = asm_mk.replace('\n\ttext_start\n\ttext_end\n', '\n\ttext_end\n')
    asm_mk = asm_mk.replace('"\n\ttext_end\n\n\ttext_end ; unused', '@"\n\ttext_end')
    asm_mk_list = asm_mk.splitlines()
    tst = False
    for line in asm_mk_list:
        if line == '':
            continue
        if tst == False:
            if 'text_start' in line:
                tst = True
        else:
            if 'para' in line or 'line' in line or 'cont' in line:
                pass
            elif 'done' in line or 'prompt' in line or 'text_end' in line:
                print('CATCH-', line, tb.olabel)
            else:
                print('ERROR-', line, tb.olabel)
            tst = False
    if asm_mk == '\tdone\n\n':
        asm_mk = '\ttext_start\n' + asm_mk
    # print(asm_mk)
    # if 'text "' not in asm_mk and 'text_start' not in asm_mk:
    #     print('STRANGE', tb.dlabel)
    #     print(''.join(tb.jplist))
    #     print('----')
    #     print(''.join(tb.cnlist))
    #     print('----')
    #     print(asm_mk)
    #     print('====')
    return asm_mk

def get_textasm(tb_dict):
    tb_asm_dict = dict()
    for tbn in tb_dict:
        tb = tb_dict[tbn]
        tb.asm = make_asm(tb)
        if tb.dlabel != '':
            if tb_asm_dict.get(tb.dmap) is None:
                tb_asm_dict[tb.dmap] = dict()
            if tb_asm_dict[tb.dmap].get(tb.dlabel) is not None:
                print('ERROR! ', tb.dlabel)
            else:
                tb_asm_dict[tb.dmap][tb.dlabel] = tb
    return tb_asm_dict

def replace_asm(asmfile_list, asmn):
    label_found_asm = set()
    opt_list = []
    state = 0
    extra_end = False
    for line in asmfile_list:
        line_strip = line[:line.find(';')].strip()
        if len(line_strip) > 0:
            if line_strip[-1] == ':':
                label = line_strip.strip(':')
                if state == 0:
                    tb = tb_asm_dict[asmn].get(label)
                elif state == 1:
                    if extra_end: extra_end = False
                        # tb.asm += '\ttext_end ; unused\n\n'
                    opt_list.append(tb.asm)
                    length_check(tb)
                    tb = tb_asm_dict[asmn].get(label)
                if tb is not None:
                    label_found_asm.add(label)
                    if len(tb.jplist) != len(tb.cnlist): trans = True
                    else:
                        trans = False
                        for i in range(len(tb.jplist)):
                            if tb.jplist[i] != tb.cnlist[i]: trans = True
                        if trans == False:
                            print('UNTRANS TEXT? ', label, tb.cnlist)
                            trans = True
                    if trans: state = 1
                    else: state = 0
                else: state = 0
            elif state == 1:
                if line == '\ttext_end ; unused\n':
                    extra_end = True
                if line[0] == '\t':
                    line = '\t; ' + line[1:]
                else:
                    line = '; ' + line
                    print(line)
        opt_list.append(line)
    if state == 1:
        opt_list.append(tb.asm)
    # print(''.join(opt_list))
    with open('./build/' + asmn, 'w') as f:
        f.writelines(opt_list)
    return label_found_asm

label_found = set()

asmfile_data = get_asmfile_data()
tb_dict = get_textdata()
tb_asm_dict = get_textasm(tb_dict)
for asmn in asmfile_data:
    asmfile = asmfile_data[asmn]
    label_found = label_found.union(replace_asm(asmfile, asmn))

# print(label_found)
for label in tb_dict:
    if tb_dict[label].dlabel not in label_found:
        # pass
        print('LABEL LEFT', 'D[ ', tb_dict[label].dlabel, ' ] O[ ', label,' ]')

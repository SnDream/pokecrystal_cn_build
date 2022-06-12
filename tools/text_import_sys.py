#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl import load_workbook

wb = load_workbook('./text.xlsx', data_only=True)

# import pokemon name
def import_pokemon():
    print('import pokemon name')
    ws = wb['宝']
    wbi = 2
    with open('./pokecrystal_cn/data/pokemon/names.asm', 'r') as f:
        file = f.readlines()
    file_opt = []
    for line in file:
        if 'db "' in line:
            name = ws.cell(row = wbi, column = 4).value
            wbi += 1
            name_len = len(name.encode(encoding = 'gb2312'))
            name_pad = name + '@' * (10 - name_len)
            line_sp = line.split('"')
            line_sp[1] = name_pad
            line = '"'.join(line_sp)
        file_opt.append(line)
    with open('./pokecrystal_cn/data/pokemon/names.asm', 'w') as f:
        f.writelines(file_opt)

# import trainer name
def import_trainer():
    print('import trainer name')
    ws = wb['训']
    wbi = 2
    with open('./pokecrystal_cn/data/trainers/parties.asm', 'r') as f:
        file = f.readlines()
    file_opt = []
    for line in file:
        if 'db "' in line and line[0] != ';':
            name = ws.cell(row = wbi, column = 4).value
            wbi += 1
            line_sp = line.split('"')
            line_sp[1] = name + '@'
            line = '"'.join(line_sp)
        file_opt.append(line)
    with open('./pokecrystal_cn/data/trainers/parties.asm', 'w') as f:
        f.writelines(file_opt)

# import class name
def import_class():
    print('import class name')
    ws = wb['类']
    wbi = 2
    with open('./pokecrystal_cn/data/trainers/class_names.asm', 'r') as f:
        file = f.readlines()
    file_opt = []
    for line in file:
        if 'db "' in line:
            name = ws.cell(row = wbi, column = 4).value
            wbi += 1
            line_sp = line.split('"')
            line_sp[1] = name + '@'
            line = '"'.join(line_sp)
        file_opt.append(line)
    with open('./pokecrystal_cn/data/trainers/class_names.asm', 'w') as f:
        f.writelines(file_opt)

# import map name
def import_map():
    print('import map name')
    ws = wb['城']
    wbi = 2
    with open('./pokecrystal_cn/data/maps/landmarks.asm', 'r') as f:
        file = f.readlines()
    file_opt = []
    for line in file:
        if 'db "' in line:
            name = ws.cell(row = wbi, column = 4).value
            wbi += 1
            line_sp = line.split('"')
            line_sp[1] = name + '@'
            line = '"'.join(line_sp)
        file_opt.append(line)
    with open('./pokecrystal_cn/data/maps/landmarks.asm', 'w') as f:
        f.writelines(file_opt)

# import item name
def import_item():
    print('import item name')
    ws = wb['道']
    wbi = 2
    with open('./pokecrystal_cn/data/items/names.asm', 'r') as f:
        file = f.readlines()
    file_opt = []
    for line in file:
        if 'db "?@"' in line:
            pass
        elif 'db "' in line:
            name = ws.cell(row = wbi, column = 4).value
            wbi += 1
            line_sp = line.split('"')
            line_sp[1] = name + '@'
            line = '"'.join(line_sp)
        file_opt.append(line)
    with open('./pokecrystal_cn/data/items/names.asm', 'w') as f:
        f.writelines(file_opt)

# import move name
def import_move():
    print('import move name')
    ws = wb['招']
    wbi = 2
    with open('./pokecrystal_cn/data/moves/names.asm', 'r') as f:
        file = f.readlines()
    file_opt = []
    for line in file:
        if 'db "' in line:
            name = ws.cell(row = wbi, column = 4).value
            wbi += 1
            line_sp = line.split('"')
            line_sp[1] = name + '@'
            line = '"'.join(line_sp)
        file_opt.append(line)
    with open('./pokecrystal_cn/data/moves/names.asm', 'w') as f:
        f.writelines(file_opt)
    
# import dex content
# skip species, height and weight
def import_dex():
    print('import dex')
    count = 0
    ws = wb ['图']
    wbi = 1
    for i in range(251):
        wbi1 = i * 9 + 1
        wbi2 = i * 9 + 6
        wbi3 = i * 9 + 7
        wbi4 = i * 9 + 8
        filename = ws.cell(row = wbi1, column = 5).value
        with open('./pokecrystal_cn/data/pokemon/dex_entries/' + filename + '.asm', 'r') as f:
            file_opt = f.readlines()
        line1 = ws.cell(row = wbi2, column = 5).value
        line2 = ws.cell(row = wbi3, column = 5).value
        line3 = ws.cell(row = wbi4, column = 5).value
        if len(line1) > 12 : print(line1)
        if len(line2) > 12 : print(line2)
        if len(line3) > 12 : print(line3)
        file_opt[3] = '\tdb   "' + line1 + '"\n'
        file_opt[4] = '\tnext "' + line2 + '"\n'
        file_opt[5] = '\tnext "' + line3 + '@"\n'
        with open('./pokecrystal_cn/data/pokemon/dex_entries/' + filename + '.asm', 'w') as f:
            f.writelines(file_opt)

import_pokemon()
import_trainer()
import_class()
import_map()
import_item()
import_move()
import_dex()
